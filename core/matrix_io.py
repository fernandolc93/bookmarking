"""
matrix_io.py
------------
Import/export helpers for the Form x Visit bookmark matrix used in Step 4
of the app. Lets users fill the matrix offline in Excel (much faster than
clicking through a 50x40-cell grid in the browser) and bring it back in.
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

TICK_VALUES = {"x", "yes", "y", "1", "true"}
_SKIP_HEADERS = {"start", "end", "status", "notes", "form_name", "form name"}


def matrix_to_excel_bytes(matrix: pd.DataFrame) -> bytes:
    """
    Serializes the current matrix to a styled .xlsx file:
      Row 1 : instructions banner
      Row 2 : headers (form_name + rotated visit names)
      Row 3+: one row per form, 'x' marking ticked cells
    Round-trips with parse_imported_matrix().
    """
    visit_cols = [c for c in matrix.columns if c != "Form Name"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form x Visit Matrix"

    # ── Styles ────────────────────────────────────────────────────────────
    ts    = Side(style="thin",   color="CCCCCC")
    ms    = Side(style="medium", color="1F4E79")
    bdr   = Border(left=ts, right=ts, top=ts, bottom=ts)
    hbdr  = Border(left=ts, right=ts, top=ts, bottom=ms)
    h_f   = Font(name="Arial", bold=True, size=9,  color="FFFFFF")
    b_f   = Font(name="Arial", size=9)
    x_f   = Font(name="Arial", bold=True, size=11, color="1F4E79")
    h_fi  = PatternFill("solid", fgColor="1F4E79")
    v_fi  = PatternFill("solid", fgColor="2E6DA4")
    y_fi  = PatternFill("solid", fgColor="FFFDE7")
    l_fi  = PatternFill("solid", fgColor="EBF3FB")
    ins_f = PatternFill("solid", fgColor="E8F5E9")

    # ── Row 1: instructions ───────────────────────────────────────────────
    last_col = get_column_letter(1 + len(visit_cols))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = (
        "INSTRUCTIONS:  Type  x  in a cell to assign that form to that visit.  "
        "Leave blank to exclude.  Column A is frozen.  "
        "Save this file when done, then import it back in the app "
        "(Step 4 → Import / Export)."
    )
    ws["A1"].font      = Font(name="Arial", italic=True, size=9, color="2E7D32")
    ws["A1"].fill      = ins_f
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: headers ────────────────────────────────────────────────────
    c = ws.cell(2, 1, "form_name")
    c.font = h_f; c.fill = h_fi
    c.alignment = Alignment(horizontal="center", vertical="bottom")
    c.border = hbdr

    for vi, visit in enumerate(visit_cols):
        col = vi + 2
        c   = ws.cell(2, col, visit)
        c.font      = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill      = v_fi
        c.alignment = Alignment(horizontal="center", vertical="bottom",
                                text_rotation=90, wrap_text=False)
        c.border    = hbdr
        ws.column_dimensions[get_column_letter(col)].width = 4.0

    ws.row_dimensions[2].height = 100

    # ── Data rows ─────────────────────────────────────────────────────────
    for fi, (_, mrow) in enumerate(matrix.iterrows()):
        rn   = fi + 3
        fill = y_fi if fi % 2 == 0 else l_fi

        ws.cell(rn, 1, str(mrow["Form Name"])).font = b_f
        ws.cell(rn, 1).alignment = Alignment(vertical="center")
        ws.cell(rn, 1).fill   = fill
        ws.cell(rn, 1).border = bdr

        for vi, visit in enumerate(visit_cols):
            col  = vi + 2
            val  = mrow[visit]
            tick = "x" if (val is True or str(val).strip().lower() in ("true", "1", "x")) else ""
            c = ws.cell(rn, col, tick)
            c.fill = fill; c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = x_f

        ws.row_dimensions[rn].height = 16

    ws.column_dimensions["A"].width = 48
    ws.freeze_panes = "B3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_imported_matrix(uploaded_file, known_forms, known_visits):
    """
    Reads an uploaded .xlsx Form x Visit matrix — either the format produced
    by matrix_to_excel_bytes() above, or the one from generate_matrix.py
    (which has an instructions row above the header) — and returns:

        imported : list[(form_name, visit_name, bool)]
            One entry per (form, visit) cell that matches a form/visit
            already present in the current session's matrix. `bool` is
            True if the cell was ticked in the file, False otherwise —
            so importing always fully overwrites the matched cells
            (lets you use Excel to clear ticks too, not just add them).
        warnings : list[str]
            Human-readable notes about anything in the file that didn't
            match the current form/visit list and was skipped.

    Raises ValueError if no 'form_name' column header can be found.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find the header row: first row (within the first 10) containing a
    # cell that reads "form_name" or "form name", case-insensitive.
    header_row = None
    max_scan_row = min(ws.max_row, 10)
    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val and str(val).strip().lower() in {"form_name", "form name"}:
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError(
            "Couldn't find a 'form_name' column in the first "
            f"{max_scan_row} rows of that file."
        )

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    form_col = next(
        i for i, h in enumerate(headers)
        if h and str(h).strip().lower() in {"form_name", "form name"}
    ) + 1

    visit_cols_in_file = [
        (i + 1, str(h).strip())
        for i, h in enumerate(headers)
        if h and str(h).strip().lower() not in _SKIP_HEADERS
    ]

    known_forms_set = set(known_forms)
    known_visits_set = set(known_visits)

    imported = []
    unmatched_forms = set()
    unmatched_visits = {v for _, v in visit_cols_in_file if v not in known_visits_set}

    for row in ws.iter_rows(min_row=header_row + 1):
        form_cell = row[form_col - 1].value
        if not form_cell:
            continue
        form_name = str(form_cell).strip()
        if form_name not in known_forms_set:
            unmatched_forms.add(form_name)
            continue
        for col_idx, visit_name in visit_cols_in_file:
            if visit_name not in known_visits_set:
                continue
            val = row[col_idx - 1].value
            ticked = val is not None and str(val).strip().lower() in TICK_VALUES
            imported.append((form_name, visit_name, ticked))

    warnings = []
    if unmatched_forms:
        names = sorted(unmatched_forms)
        shown = ", ".join(names[:8]) + (" …" if len(names) > 8 else "")
        warnings.append(
            f"{len(unmatched_forms)} form(s) in the file don't match your "
            f"current form list and were skipped: {shown}"
        )
    if unmatched_visits:
        names = sorted(unmatched_visits)
        shown = ", ".join(names[:8]) + (" …" if len(names) > 8 else "")
        warnings.append(
            f"{len(unmatched_visits)} visit column(s) in the file don't match "
            f"your current visit list and were skipped: {shown}"
        )

    return imported, warnings