"""
generate_matrix.py
------------------
Reads bookmark_template.xlsx (tabs: Form Mapping + Schedule of Assessments)
and generates bookmark_matrix.xlsx — a Form x Visit cross-table where you
type  x  to assign a form to a visit.

Usage:
    python generate_matrix.py

Input  : acrf_bookmarking/bookmark_template.xlsx   (already filled)
Output : acrf_bookmarking/bookmark_matrix.xlsx      (open and tick)
"""

import sys, subprocess

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

TEMPLATE_PATH = Path("acrf_bookmarking/bookmark_template.xlsx")
MATRIX_PATH   = Path("acrf_bookmarking/bookmark_matrix.xlsx")

assert TEMPLATE_PATH.exists(), f"Not found: {TEMPLATE_PATH}"

# ── Read Form Mapping tab ──────────────────────────────────────────────────────
wb_in    = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
ws_forms = wb_in["Form Mapping"]
forms    = []
for row in ws_forms.iter_rows(min_row=3, values_only=True):
    name, start, end, count, status, notes = (list(row) + [None]*6)[:6]
    if not name:
        continue
    if str(status or "").strip().lower() == "not submitted":
        continue
    forms.append({
        "form_name"  : str(name).strip(),
        "start_page" : int(start) if start else 0,
        "end_page"   : int(end)   if end   else 0,
    })

# ── Read Schedule of Assessments tab ──────────────────────────────────────────
ws_soa = wb_in["Schedule of Assessments"]
visits = []
for row in ws_soa.iter_rows(min_row=3, values_only=True):
    num, name, notes = (list(row) + [None]*3)[:3]
    if name:
        visits.append(str(name).strip())

print(f"Forms loaded  : {len(forms)}  (excluded 'Not submitted')")
print(f"Visits loaded : {len(visits)}")

# ── Build matrix workbook ─────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws     = wb_out.active
ws.title = "Form × Visit Matrix"

ts    = Side(style="thin",   color="CCCCCC")
ms    = Side(style="medium", color="1F4E79")
bdr   = Border(left=ts, right=ts, top=ts, bottom=ts)
hbdr  = Border(left=ts, right=ts, top=ts, bottom=ms)
h_f   = Font(name="Arial", bold=True, size=9,  color="FFFFFF")
b_f   = Font(name="Arial", size=9)
sm_f  = Font(name="Arial", size=8,  color="555555")
x_f   = Font(name="Arial", bold=True, size=11, color="1F4E79")
h_fi  = PatternFill("solid", fgColor="1F4E79")
v_fi  = PatternFill("solid", fgColor="2E6DA4")
y_fi  = PatternFill("solid", fgColor="FFFDE7")
l_fi  = PatternFill("solid", fgColor="EBF3FB")
ins_f = PatternFill("solid", fgColor="E8F5E9")

# Row 1: instructions
last_col = get_column_letter(3 + len(visits))
ws.merge_cells(f"A1:{last_col}1")
ws["A1"] = (
    "INSTRUCTIONS:  Type  x  in a cell to assign that form to that visit.  "
    "Leave blank to exclude.  Columns A-C are frozen.  "
    "Save this file when done, then run notebook 02."
)
ws["A1"].font      = Font(name="Arial", italic=True, size=9, color="2E7D32")
ws["A1"].fill      = ins_f
ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[1].height = 36

# Row 2: headers
for col, text in enumerate(["start", "end", "form_name"], 1):
    c = ws.cell(2, col, text)
    c.font = h_f; c.fill = h_fi
    c.alignment = Alignment(horizontal="center", vertical="bottom")
    c.border = hbdr

for vi, visit in enumerate(visits):
    col = vi + 4
    c   = ws.cell(2, col, visit)
    c.font      = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    c.fill      = v_fi
    c.alignment = Alignment(horizontal="center", vertical="bottom",
                            text_rotation=90, wrap_text=False)
    c.border    = hbdr
    ws.column_dimensions[get_column_letter(col)].width = 4.0

ws.row_dimensions[2].height = 100

# Data rows
for fi, form in enumerate(forms):
    rn   = fi + 3
    fill = y_fi if fi % 2 == 0 else l_fi

    ws.cell(rn, 1, form["start_page"]).font = sm_f
    ws.cell(rn, 2, form["end_page"]).font   = sm_f
    ws.cell(rn, 3, form["form_name"]).font  = b_f
    ws.cell(rn, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(rn, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(rn, 3).alignment = Alignment(vertical="center")

    for vi in range(len(visits)):
        col = vi + 4
        c   = ws.cell(rn, col, "")
        c.fill = fill; c.border = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = x_f

    for col in [1, 2, 3]:
        ws.cell(rn, col).fill   = fill
        ws.cell(rn, col).border = bdr

    ws.row_dimensions[rn].height = 16

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 48
ws.freeze_panes = "D3"

wb_out.save(MATRIX_PATH)
print(f"\nSaved: {MATRIX_PATH}")
print(f"  {len(forms)} forms  x  {len(visits)} visits  =  {len(forms)*len(visits)} cells")
print(f"\nNext: open {MATRIX_PATH.name}, type x to assign forms to visits, then run notebook 02.")

